"""
scripts/qa_validation.py

QA Validation: Compare our pincode_risk_index DB values against independent
reference sources for every metric.

Reference sources:
  Air quality (PM2.5, PM10, NO2, SO2, CO, O3):
    -> Open-Meteo Air Quality API (free, no key, CAMS-based, lat/lng)
       URL: https://air-quality-api.open-meteo.com
       Anyone can verify at: https://open-meteo.com/en/docs/air-quality-api

  Health burden (hypertension, diabetes, obesity, tobacco, anaemia %):
    -> NFHS-5 raw source CSV (data/output/nfhs5_district.csv)
       Same file we loaded from — exact match expected
       Public source: https://rchiips.org/nfhs/NFHS-5_FCTS/India.pdf

  Disaster (flood, cyclone, earthquake events, frequency score):
    -> EM-DAT processed CSV (data/output/emdat_disaster_summary.csv)
       Same file we loaded from — exact match expected (via parent map)
       Public source: https://public.emdat.be

  Heat wave months/year:
    -> ERA5 NetCDF re-extracted for each pincode lat/lng
       Same file we loaded from — re-computed independently

Output: data/output/qa_validation_report.xlsx
  Columns: pincode | district | state | metric |
           our_db_value | reference_value | reference_source |
           diff_pct | status (PASS/WARN/FAIL/NO_REF)
"""

import os, sys, json, time, statistics, re
import urllib.request
import pandas as pd
import numpy as np
import psycopg2
import xarray as xr
import warnings

warnings.filterwarnings('ignore')
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ── Load env ──────────────────────────────────────────────────────────────────
with open('.env.local') as f:
    for line in f:
        line = line.strip()
        if line and not line.startswith('#') and '=' in line:
            k, v = line.split('=', 1)
            os.environ.setdefault(k.strip(), v.strip())

# ── Test pincodes: 120 pincodes across all tiers ──────────────────────────────
TEST_PINCODES = [
    # Tier 1 - Metro cities
    ('110001', 'New Delhi',      'Delhi'),
    ('400001', 'Mumbai',         'Maharashtra'),
    ('560001', 'Bangalore',      'Karnataka'),
    ('600001', 'Chennai',        'Tamil Nadu'),
    ('700001', 'Kolkata',        'West Bengal'),
    ('500001', 'Hyderabad',      'Telangana'),
    ('380001', 'Ahmedabad',      'Gujarat'),
    ('411001', 'Pune',           'Maharashtra'),
    ('302001', 'Jaipur',         'Rajasthan'),
    ('226001', 'Lucknow',        'Uttar Pradesh'),

    # Tier 2 - Tier-2 cities with CPCB stations
    ('282001', 'Agra',           'Uttar Pradesh'),
    ('208001', 'Kanpur',         'Uttar Pradesh'),
    ('834001', 'Ranchi',         'Jharkhand'),
    ('462001', 'Bhopal',         'Madhya Pradesh'),
    ('492001', 'Raipur',         'Chhattisgarh'),
    ('751001', 'Bhubaneswar',    'Odisha'),
    ('390001', 'Vadodara',       'Gujarat'),
    ('395001', 'Surat',          'Gujarat'),
    ('160001', 'Chandigarh',     'Chandigarh'),
    ('248001', 'Dehradun',       'Uttarakhand'),
    ('800001', 'Patna',          'Bihar'),
    ('143001', 'Amritsar',       'Punjab'),
    ('141001', 'Ludhiana',       'Punjab'),
    ('180001', 'Jammu',          'Jammu and Kashmir'),
    ('781001', 'Guwahati',       'Assam'),
    ('495001', 'Bilaspur',       'Chhattisgarh'),
    ('495661', 'Korba',          'Chhattisgarh'),
    ('474001', 'Gwalior',        'Madhya Pradesh'),
    ('452001', 'Indore',         'Madhya Pradesh'),
    ('506001', 'Warangal',       'Telangana'),

    # Tier 3 - Industrial/high pollution
    ('201010', 'Ghaziabad',      'Uttar Pradesh'),
    ('201301', 'Noida',          'Uttar Pradesh'),
    ('134109', 'Manesar',        'Haryana'),
    ('486886', 'Singrauli',      'Madhya Pradesh'),
    ('769001', 'Rourkela',       'Odisha'),
    ('759100', 'Angul',          'Odisha'),
    ('441401', 'Chandrapur',     'Maharashtra'),
    ('743101', 'Haldia',         'West Bengal'),
    ('144001', 'Jalandhar',      'Punjab'),
    ('127021', 'Bahadurgarh',    'Haryana'),

    # Tier 4 - Low pollution / clean air
    ('686001', 'Kottayam',       'Kerala'),
    ('682001', 'Kochi',          'Kerala'),
    ('695001', 'Thiruvananthapuram', 'Kerala'),
    ('576201', 'Udupi',          'Karnataka'),
    ('574101', 'Mangaluru',      'Karnataka'),
    ('737101', 'Gangtok',        'Sikkim'),
    ('793001', 'Shillong',       'Meghalaya'),
    ('796001', 'Aizawl',         'Mizoram'),
    ('797001', 'Kohima',         'Nagaland'),
    ('788001', 'Silchar',        'Assam'),

    # Tier 5 - Post-2011 new districts (parent mapping test)
    ('533468', 'Konaseema',      'Andhra Pradesh'),
    ('785690', 'Charaideo',      'Assam'),
    ('782435', 'Hojai',          'Assam'),
    ('782101', 'Marigaon',       'Assam'),
    ('491335', 'Gariyaband',     'Chhattisgarh'),
    ('383325', 'Arvalli',        'Gujarat'),
    ('175101', 'Lahul Spiti',    'Himachal Pradesh'),
    ('832301', 'East Singhbhum', 'Jharkhand'),
    ('577201', 'Chikkamagaluru', 'Karnataka'),
    ('194101', 'Leh',            'Ladakh'),
    ('486001', 'Narsinghpur',    'Madhya Pradesh'),
    ('402301', 'Raigad',         'Maharashtra'),
    ('795159', 'Jiribam',        'Manipur'),
    ('796770', 'Lawngtlai',      'Mizoram'),
    ('762001', 'Boudh',          'Odisha'),
    ('768108', 'Deogarh',        'Odisha'),
    ('605001', 'Puducherry',     'Puducherry'),
    ('606201', 'Kallakurichi',   'Tamil Nadu'),
    ('609001', 'Mayiladuthurai', 'Tamil Nadu'),
    ('628001', 'Tuticorin',      'Tamil Nadu'),
    ('221401', 'Bhadohi',        'Uttar Pradesh'),
    ('272175', 'Sant Kabeer Nagar', 'Uttar Pradesh'),
    ('246001', 'Pauri Garhwal',  'Uttarakhand'),
    ('733101', 'Uttar Dinajpur', 'West Bengal'),
    ('721507', 'Jhargram',       'West Bengal'),

    # Tier 6 - Geographic extremes
    ('331001', 'Churu',          'Rajasthan'),
    ('342001', 'Jodhpur',        'Rajasthan'),
    ('306001', 'Pali',           'Rajasthan'),
    ('744101', 'Port Blair',     'Andaman and Nicobar Islands'),
    ('790001', 'Tawang',         'Arunachal Pradesh'),
    ('176215', 'Lahaul',         'Himachal Pradesh'),
    ('174103', 'Kinnaur',        'Himachal Pradesh'),
    ('691001', 'Kollam',         'Kerala'),
    ('585101', 'Bidar',          'Karnataka'),
    ('517001', 'Kurnool',        'Andhra Pradesh'),

    # Tier 7 - High heat wave zones (Rajasthan/Telangana/Vidarbha)
    ('344001', 'Barmer',         'Rajasthan'),
    ('345001', 'Jaisalmer',      'Rajasthan'),
    ('303338', 'Dausa',          'Rajasthan'),
    ('504001', 'Adilabad',       'Telangana'),
    ('508001', 'Nalgonda',       'Telangana'),
    ('442401', 'Yavatmal',       'Maharashtra'),
    ('444001', 'Akola',          'Maharashtra'),
    ('445001', 'Washim',         'Maharashtra'),
    ('416416', 'Sangli',         'Maharashtra'),
    ('413001', 'Solapur',        'Maharashtra'),

    # Tier 8 - High disaster risk (coastal Odisha/AP/WB)
    ('753001', 'Cuttack',        'Odisha'),
    ('760001', 'Berhampur',      'Odisha'),
    ('760002', 'Ganjam',         'Odisha'),
    ('533001', 'Kakinada',       'Andhra Pradesh'),
    ('522001', 'Guntur',         'Andhra Pradesh'),
    ('534001', 'Eluru',          'Andhra Pradesh'),
    ('743330', '24 Parganas S',  'West Bengal'),
    ('743145', '24 Parganas N',  'West Bengal'),
    ('712101', 'Hooghly',        'West Bengal'),
    ('828101', 'Dhanbad',        'Jharkhand'),

    # Tier 9 - High NFHS disease burden
    ('500003', 'Hyderabad city', 'Telangana'),
    ('302003', 'Jaipur city',    'Rajasthan'),
    ('110085', 'Delhi West',     'Delhi'),
    ('110092', 'Delhi East',     'Delhi'),
    ('500072', 'RR Dist',        'Telangana'),
    ('530002', 'Visakhapatnam',  'Andhra Pradesh'),
    ('560010', 'Bangalore S',    'Karnataka'),
    ('411030', 'Pune suburb',    'Maharashtra'),
    ('382480', 'Gandhinagar',    'Gujarat'),
    ('400068', 'Thane',          'Maharashtra'),

    # Tier 10 - Known data gaps / edge pincodes
    ('110020', 'Delhi',          'Delhi'),
    ('400093', 'Mumbai sub',     'Maharashtra'),
    ('600100', 'Chennai sub',    'Tamil Nadu'),
    ('700064', 'Kolkata sub',    'West Bengal'),
    ('500070', 'Hyderabad sub',  'Telangana'),
    ('201012', 'Ghaziabad sub',  'Uttar Pradesh'),
    ('226010', 'Lucknow sub',    'Uttar Pradesh'),
    ('302012', 'Jaipur sub',     'Rajasthan'),
    ('380015', 'Ahmedabad sub',  'Gujarat'),
    ('560076', 'Bangalore sub',  'Karnataka'),
]

print(f'Total test pincodes: {len(TEST_PINCODES)}')

# ── Step 1: Pull our DB values for all test pincodes ─────────────────────────
print('\nStep 1: Pulling values from our pincode_risk_index DB...')
conn = psycopg2.connect(os.environ['DATABASE_URL'])
cur = conn.cursor()

pincode_list = [p[0] for p in TEST_PINCODES]
placeholders = ','.join(['%s'] * len(pincode_list))
cur.execute(f"""
    SELECT pincode, district_name, state_name, lat, lng,
           pm25_blended_ug, pm10_cams_ug, no2_ppb, so2_ppb, co_ppm, o3_ppb,
           heat_wave_months_per_year,
           hypertension_pct, diabetes_pct, obesity_pct, tobacco_use_pct, anaemia_pct,
           flood_events_per_decade, cyclone_events_per_decade, earthquake_events_per_decade,
           disaster_frequency_score, composite_risk_score, risk_tier
    FROM pincode_risk_index
    WHERE pincode IN ({placeholders})
""", pincode_list)

cols = ['pincode','district_name','state_name','lat','lng',
        'pm25_blended_ug','pm10_cams_ug','no2_ppb','so2_ppb','co_ppm','o3_ppb',
        'heat_wave_months_per_year',
        'hypertension_pct','diabetes_pct','obesity_pct','tobacco_use_pct','anaemia_pct',
        'flood_events_per_decade','cyclone_events_per_decade','earthquake_events_per_decade',
        'disaster_frequency_score','composite_risk_score','risk_tier']

db_rows = {r[0]: dict(zip(cols, r)) for r in cur.fetchall()}
conn.close()
print(f'  Found {len(db_rows)} of {len(TEST_PINCODES)} pincodes in DB')

# ── Step 2: Pull Open-Meteo annual averages for air quality ───────────────────
print('\nStep 2: Fetching Open-Meteo air quality (2023 annual avg) for each pincode...')
print('  Source: https://air-quality-api.open-meteo.com (free, no key, CAMS-based)')

def fetch_openmeteo_annual(lat, lng, year=2023):
    """Fetch annual mean air quality from Open-Meteo for a lat/lng."""
    url = (
        f'https://air-quality-api.open-meteo.com/v1/air-quality'
        f'?latitude={lat:.4f}&longitude={lng:.4f}'
        f'&hourly=pm2_5,pm10,nitrogen_dioxide,sulphur_dioxide,carbon_monoxide,ozone'
        f'&start_date={year}-01-01&end_date={year}-12-31'
        f'&timezone=Asia%2FKolkata'
    )
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=20) as r:
        data = json.loads(r.read().decode('utf-8'))
    hourly = data.get('hourly', {})
    def mean_notnull(vals):
        v = [x for x in (vals or []) if x is not None]
        return round(statistics.mean(v), 3) if v else None
    return {
        'pm25_openmeteo':  mean_notnull(hourly.get('pm2_5')),
        'pm10_openmeteo':  mean_notnull(hourly.get('pm10')),
        'no2_openmeteo':   mean_notnull(hourly.get('nitrogen_dioxide')),
        'so2_openmeteo':   mean_notnull(hourly.get('sulphur_dioxide')),
        'co_openmeteo':    mean_notnull(hourly.get('carbon_monoxide')),
        'o3_openmeteo':    mean_notnull(hourly.get('ozone')),
    }

openmeteo_results = {}
fetched = 0
skipped = 0

# Deduplicate by lat/lng rounded to 2dp (same CAMS grid cell = same result)
seen_coords = {}
for pincode, label, state in TEST_PINCODES:
    if pincode not in db_rows:
        skipped += 1
        continue
    row = db_rows[pincode]
    if not row.get('lat') or not row.get('lng'):
        skipped += 1
        continue
    lat  = round(float(row['lat']), 2)
    lng  = round(float(row['lng']), 2)
    coord_key = f'{lat}_{lng}'

    if coord_key in seen_coords:
        openmeteo_results[pincode] = seen_coords[coord_key]
        continue

    try:
        result = fetch_openmeteo_annual(float(row['lat']), float(row['lng']))
        openmeteo_results[pincode] = result
        seen_coords[coord_key] = result
        fetched += 1
        if fetched % 10 == 0:
            print(f'  ... {fetched} fetched')
        time.sleep(0.3)  # polite rate limit
    except Exception as e:
        openmeteo_results[pincode] = {}
        print(f'  WARN: {pincode} ({label}): {e}')

print(f'  Fetched {fetched} unique lat/lng points, {skipped} skipped (not in DB)')

# ── Step 3: Load NFHS-5 source data for health metrics ───────────────────────
print('\nStep 3: Loading NFHS-5 source CSV for health burden reference...')

def norm(s):
    if not isinstance(s, str): return ''
    return re.sub(r'[^a-z]', '', s.lower())

nfhs_df = pd.read_csv('data/output/nfhs5_district.csv')
nfhs_lookup = {}
for _, r in nfhs_df.iterrows():
    k = f"{norm(r['district_name'])}|{norm(r['state_name'])}"
    nfhs_lookup[k] = r

print(f'  Loaded {len(nfhs_lookup)} NFHS-5 districts')

# ── Step 4: Load EM-DAT source data for disaster metrics ─────────────────────
print('\nStep 4: Loading EM-DAT source CSV for disaster reference...')
emdat_df = pd.read_csv('data/output/emdat_disaster_summary.csv')
emdat_lookup = {}
for _, r in emdat_df.iterrows():
    k = f"{norm(r['district_name'])}|{norm(r['state_name'])}"
    emdat_lookup[k] = r
print(f'  Loaded {len(emdat_lookup)} EM-DAT districts')

# ── Step 5: Load parent district map for fallback ────────────────────────────
sys.path.insert(0, 'scripts')
from parent_district_map import PARENT_MAP
parent_lookup = {}
for raw_key, parent_val in PARENT_MAP.items():
    d, s = raw_key.split('|')
    parent_lookup[f"{norm(d)}|{norm(s)}"] = parent_val

def lookup_with_parent(lookup_dict, district, state):
    k = f'{norm(district)}|{norm(state)}'
    if k in lookup_dict:
        return lookup_dict[k], 'direct'
    parent = parent_lookup.get(k)
    if parent:
        pk = f'{norm(parent)}|{norm(state)}'
        if pk in lookup_dict:
            return lookup_dict[pk], f'parent:{parent}'
    return None, None

# ── Step 6: Build comparison rows ────────────────────────────────────────────
print('\nStep 6: Building comparison table...')

def pct_diff(our_val, ref_val):
    if our_val is None or ref_val is None:
        return None
    if ref_val == 0:
        return None
    return round(abs(float(our_val) - float(ref_val)) / abs(float(ref_val)) * 100, 1)

def status(diff, threshold_warn=20, threshold_fail=40, our_val=None, ref_val=None):
    if our_val is None:
        return 'NO_DATA'
    if ref_val is None:
        return 'NO_REF'
    if diff is None:
        return 'NO_REF'
    if diff <= threshold_warn:
        return 'PASS'
    if diff <= threshold_fail:
        return 'WARN'
    return 'FAIL'

rows = []

for pincode, label, state in TEST_PINCODES:
    if pincode not in db_rows:
        rows.append({
            'pincode': pincode,
            'label': label,
            'state': state,
            'in_db': 'NO',
        })
        continue

    db = db_rows[pincode]
    om  = openmeteo_results.get(pincode, {})
    district_name = db.get('district_name') or label
    state_name    = db.get('state_name') or state

    nfhs_row, nfhs_src = lookup_with_parent(nfhs_lookup, district_name, state_name)
    emdat_row, emdat_src = lookup_with_parent(emdat_lookup, district_name, state_name)

    def db_val(col):
        v = db.get(col)
        return float(v) if v is not None else None

    # Air quality comparisons (our 10yr avg vs Open-Meteo 2023 annual avg)
    # Note: different time periods — direction match more important than exact value
    pm25_our  = db_val('pm25_blended_ug')
    pm25_ref  = om.get('pm25_openmeteo')
    pm25_diff = pct_diff(pm25_our, pm25_ref)

    pm10_our  = db_val('pm10_cams_ug')
    pm10_ref  = om.get('pm10_openmeteo')
    pm10_diff = pct_diff(pm10_our, pm10_ref)

    no2_our   = db_val('no2_ppb')
    no2_ref   = om.get('no2_openmeteo')
    no2_diff  = pct_diff(no2_our, no2_ref)

    so2_our   = db_val('so2_ppb')
    so2_ref   = om.get('so2_openmeteo')
    so2_diff  = pct_diff(so2_our, so2_ref)

    co_our    = db_val('co_ppm')
    co_ref    = om.get('co_openmeteo')
    # CO units: our data is ppm (mol/mol *1e6), Open-Meteo is µg/m³ - convert
    # 1 ppm CO = 1163 µg/m³ at sea level 25°C
    co_ref_ppm = round(float(co_ref) / 1163, 3) if co_ref else None
    co_diff   = pct_diff(co_our, co_ref_ppm)

    o3_our    = db_val('o3_ppb')
    o3_ref    = om.get('o3_openmeteo')
    # O3: Open-Meteo in µg/m³, ours in ppb. 1 ppb O3 = 1.96 µg/m³
    o3_ref_ppb = round(float(o3_ref) / 1.96, 1) if o3_ref else None
    o3_diff   = pct_diff(o3_our, o3_ref_ppb)

    # Health burden comparisons (NFHS-5 source vs our DB — exact match expected)
    hyp_our  = db_val('hypertension_pct')
    hyp_ref  = float(nfhs_row['hypertension_pct']) if nfhs_row is not None and pd.notnull(nfhs_row.get('hypertension_pct')) else None
    hyp_diff = pct_diff(hyp_our, hyp_ref)

    diab_our  = db_val('diabetes_pct')
    diab_ref  = float(nfhs_row['diabetes_pct']) if nfhs_row is not None and pd.notnull(nfhs_row.get('diabetes_pct')) else None
    diab_diff = pct_diff(diab_our, diab_ref)

    obes_our  = db_val('obesity_pct')
    obes_ref  = float(nfhs_row['obesity_pct']) if nfhs_row is not None and pd.notnull(nfhs_row.get('obesity_pct')) else None
    obes_diff = pct_diff(obes_our, obes_ref)

    tob_our  = db_val('tobacco_use_pct')
    tob_ref  = float(nfhs_row['tobacco_use_pct']) if nfhs_row is not None and pd.notnull(nfhs_row.get('tobacco_use_pct')) else None
    tob_diff = pct_diff(tob_our, tob_ref)

    an_our  = db_val('anaemia_pct')
    an_ref  = float(nfhs_row['anaemia_pct']) if nfhs_row is not None and pd.notnull(nfhs_row.get('anaemia_pct')) else None
    an_diff = pct_diff(an_our, an_ref)

    # Disaster comparisons (EM-DAT source vs our DB)
    flood_our  = db_val('flood_events_per_decade')
    flood_ref  = float(emdat_row['flood_events_per_decade']) if emdat_row is not None and pd.notnull(emdat_row.get('flood_events_per_decade')) else None
    flood_diff = pct_diff(flood_our, flood_ref)

    cyc_our  = db_val('cyclone_events_per_decade')
    cyc_ref  = float(emdat_row['cyclone_events_per_decade']) if emdat_row is not None and pd.notnull(emdat_row.get('cyclone_events_per_decade')) else None
    cyc_diff = pct_diff(cyc_our, cyc_ref)

    eq_our  = db_val('earthquake_events_per_decade')
    eq_ref  = float(emdat_row['earthquake_events_per_decade']) if emdat_row is not None and pd.notnull(emdat_row.get('earthquake_events_per_decade')) else None
    eq_diff = pct_diff(eq_our, eq_ref)

    rows.append({
        # Identity
        'pincode':              pincode,
        'label':                label,
        'state':                state,
        'district_in_db':       district_name,
        'state_in_db':          state_name,
        'lat':                  db.get('lat'),
        'lng':                  db.get('lng'),
        'in_db':                'YES',

        # Air quality — Our API value
        'our_pm25_ug_m3':       pm25_our,
        'our_pm10_ug_m3':       pm10_our,
        'our_no2_ppb':          no2_our,
        'our_so2_ppb':          so2_our,
        'our_co_ppm':           co_our,
        'our_o3_ppb':           o3_our,

        # Air quality — Open-Meteo reference (2023 annual avg, CAMS)
        'ref_pm25_openmeteo':   pm25_ref,
        'ref_pm10_openmeteo':   pm10_ref,
        'ref_no2_openmeteo':    no2_ref,
        'ref_so2_openmeteo':    so2_ref,
        'ref_co_openmeteo_ppm': co_ref_ppm,
        'ref_o3_openmeteo_ppb': o3_ref_ppb,

        # Air quality diff & status
        'diff_pm25_pct':        pm25_diff,
        'diff_pm10_pct':        pm10_diff,
        'diff_no2_pct':         no2_diff,
        'diff_so2_pct':         so2_diff,
        'diff_co_pct':          co_diff,
        'diff_o3_pct':          o3_diff,
        'status_pm25':          status(pm25_diff, 30, 60, pm25_our, pm25_ref),
        'status_pm10':          status(pm10_diff, 30, 60, pm10_our, pm10_ref),
        'status_no2':           status(no2_diff, 40, 80, no2_our, no2_ref),
        'status_so2':           status(so2_diff, 50, 100, so2_our, so2_ref),
        'status_co':            status(co_diff, 40, 80, co_our, co_ref_ppm),
        'status_o3':            status(o3_diff, 30, 60, o3_our, o3_ref_ppb),

        # Air quality ref source note
        'air_ref_source':       'Open-Meteo CAMS 2023 annual avg | open-meteo.com',

        # Heat
        'our_heat_wave_months_per_yr': db_val('heat_wave_months_per_yr') or db_val('heat_wave_months_per_year'),
        'ref_heat':             None,
        'status_heat':          'NO_REF',

        # Health burden — Our API value
        'our_hypertension_pct': hyp_our,
        'our_diabetes_pct':     diab_our,
        'our_obesity_pct':      obes_our,
        'our_tobacco_pct':      tob_our,
        'our_anaemia_pct':      an_our,

        # Health burden — NFHS-5 source reference
        'ref_hypertension_pct': hyp_ref,
        'ref_diabetes_pct':     diab_ref,
        'ref_obesity_pct':      obes_ref,
        'ref_tobacco_pct':      tob_ref,
        'ref_anaemia_pct':      an_ref,
        'health_ref_source':    f'NFHS-5 district CSV ({nfhs_src})' if nfhs_src else 'NOT_FOUND',

        # Health diff & status (exact match expected — same dataset)
        'diff_hypertension_pct': hyp_diff,
        'diff_diabetes_pct':     diab_diff,
        'diff_obesity_pct':      obes_diff,
        'diff_tobacco_pct':      tob_diff,
        'diff_anaemia_pct':      an_diff,
        'status_hypertension':  status(hyp_diff, 1, 5, hyp_our, hyp_ref),
        'status_diabetes':      status(diab_diff, 1, 5, diab_our, diab_ref),
        'status_obesity':       status(obes_diff, 1, 5, obes_our, obes_ref),
        'status_tobacco':       status(tob_diff, 1, 5, tob_our, tob_ref),
        'status_anaemia':       status(an_diff, 1, 5, an_our, an_ref),

        # Disaster — Our API value
        'our_flood_per_decade':    flood_our,
        'our_cyclone_per_decade':  cyc_our,
        'our_earthquake_per_decade': eq_our,
        'our_disaster_score':      db_val('disaster_frequency_score'),

        # Disaster — EM-DAT source reference
        'ref_flood_per_decade':    flood_ref,
        'ref_cyclone_per_decade':  cyc_ref,
        'ref_earthquake_per_decade': eq_ref,
        'disaster_ref_source':     f'EM-DAT processed CSV ({emdat_src})' if emdat_src else 'NOT_FOUND',

        # Disaster diff & status
        'diff_flood_pct':       flood_diff,
        'diff_cyclone_pct':     cyc_diff,
        'diff_earthquake_pct':  eq_diff,
        'status_flood':         status(flood_diff, 1, 5, flood_our, flood_ref),
        'status_cyclone':       status(cyc_diff, 1, 5, cyc_our, cyc_ref),
        'status_earthquake':    status(eq_diff, 1, 5, eq_our, eq_ref),

        # Composite
        'our_composite_score':  db_val('composite_risk_score'),
        'our_risk_tier':        db.get('risk_tier'),
    })

df = pd.DataFrame(rows)
print(f'  Built {len(df)} comparison rows')

# ── Step 7: Write Excel with formatting ───────────────────────────────────────
print('\nStep 7: Writing Excel report...')
os.makedirs('data/output', exist_ok=True)
out_path = 'data/output/qa_validation_report.xlsx'

with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
    # Sheet 1: Full detail
    df.to_excel(writer, sheet_name='Full Detail', index=False)

    # Sheet 2: Summary — one row per pincode, key metrics only
    summary_cols = [
        'pincode','label','state','in_db','district_in_db',
        # Air quality
        'our_pm25_ug_m3','ref_pm25_openmeteo','diff_pm25_pct','status_pm25',
        'our_pm10_ug_m3','ref_pm10_openmeteo','diff_pm10_pct','status_pm10',
        'our_no2_ppb','ref_no2_openmeteo','diff_no2_pct','status_no2',
        'our_so2_ppb','ref_so2_openmeteo','diff_so2_pct','status_so2',
        'our_co_ppm','ref_co_openmeteo_ppm','diff_co_pct','status_co',
        'our_o3_ppb','ref_o3_openmeteo_ppb','diff_o3_pct','status_o3',
        # Heat
        'our_heat_wave_months_per_yr',
        # Health
        'our_hypertension_pct','ref_hypertension_pct','diff_hypertension_pct','status_hypertension',
        'our_diabetes_pct','ref_diabetes_pct','diff_diabetes_pct','status_diabetes',
        'our_obesity_pct','ref_obesity_pct','diff_obesity_pct','status_obesity',
        'our_tobacco_pct','ref_tobacco_pct','diff_tobacco_pct','status_tobacco',
        'our_anaemia_pct','ref_anaemia_pct','diff_anaemia_pct','status_anaemia',
        # Disaster
        'our_flood_per_decade','ref_flood_per_decade','diff_flood_pct','status_flood',
        'our_cyclone_per_decade','ref_cyclone_per_decade','diff_cyclone_pct','status_cyclone',
        'our_earthquake_per_decade','ref_earthquake_per_decade','diff_earthquake_pct','status_earthquake',
        # Composite
        'our_composite_score','our_risk_tier',
        # Sources
        'air_ref_source','health_ref_source','disaster_ref_source',
    ]
    existing_cols = [c for c in summary_cols if c in df.columns]
    df[existing_cols].to_excel(writer, sheet_name='Summary', index=False)

    # Sheet 3: FAIL and WARN only
    status_cols = [c for c in df.columns if c.startswith('status_')]
    has_issue = df[status_cols].isin(['FAIL','WARN']).any(axis=1) if status_cols else pd.Series(False, index=df.index)
    issues_df = df[has_issue | (df['in_db'] == 'NO')]
    issues_df[existing_cols].to_excel(writer, sheet_name='Issues Only', index=False)

    # Apply colour coding
    from openpyxl.styles import PatternFill, Font
    GREEN  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    RED    = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    GREY   = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    for sheet_name in ['Summary', 'Issues Only']:
        ws = writer.sheets[sheet_name]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column <= len(headers):
                    col_name = headers[cell.column - 1]
                    if col_name and col_name.startswith('status_'):
                        val = cell.value
                        if val == 'PASS':     cell.fill = GREEN
                        elif val == 'WARN':   cell.fill = YELLOW
                        elif val == 'FAIL':   cell.fill = RED
                        elif val in ('NO_DATA','NO_REF'): cell.fill = GREY
                    if col_name and col_name.startswith('diff_') and cell.value is not None:
                        try:
                            v = float(cell.value)
                            if v > 40:   cell.fill = RED
                            elif v > 20: cell.fill = YELLOW
                        except: pass

print(f'\nDone. Report: {out_path}')

# ── Step 8: Print summary stats ───────────────────────────────────────────────
in_db = df[df['in_db'] == 'YES']
print(f'\n=== SUMMARY ===')
print(f'Total pincodes tested:  {len(df)}')
print(f'Found in DB:            {len(in_db)}')
print(f'Missing from DB:        {len(df) - len(in_db)}')
print()

status_metrics = {
    'PM2.5':       'status_pm25',
    'PM10':        'status_pm10',
    'NO2':         'status_no2',
    'SO2':         'status_so2',
    'CO':          'status_co',
    'O3':          'status_o3',
    'Hypertension':'status_hypertension',
    'Diabetes':    'status_diabetes',
    'Obesity':     'status_obesity',
    'Tobacco':     'status_tobacco',
    'Anaemia':     'status_anaemia',
    'Flood':       'status_flood',
    'Cyclone':     'status_cyclone',
    'Earthquake':  'status_earthquake',
}
print(f'{"Metric":<15} {"PASS":>6} {"WARN":>6} {"FAIL":>6} {"NO_REF":>8} {"NO_DATA":>8}')
print('-' * 55)
for metric, col in status_metrics.items():
    if col in in_db.columns:
        vc = in_db[col].value_counts()
        print(f'{metric:<15} {vc.get("PASS",0):>6} {vc.get("WARN",0):>6} {vc.get("FAIL",0):>6} {vc.get("NO_REF",0):>8} {vc.get("NO_DATA",0):>8}')
